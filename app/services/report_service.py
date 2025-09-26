import os
import uuid
import tempfile
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional, Union
from pathlib import Path
import logging

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from sqlalchemy.orm import Session
from sqlalchemy import text

from app.models.reports import GeneratedReport, ReportTemplate, ReportAccessLog
from app.schemas.reports import (
    ReportGenerationRequest, ReportFormat, ReportStatus,
    ClinicalReportParameters, FinancialReportParameters,
    CommercialReportParameters, AdministrativeReportParameters
)

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Base class for report generation"""
    
    def __init__(self, db: Session):
        self.db = db
        self.temp_dir = Path(tempfile.gettempdir()) / "prontivus_reports"
        self.temp_dir.mkdir(exist_ok=True)
    
    def generate_report(self, request: ReportGenerationRequest, user_id: int) -> GeneratedReport:
        """Generate a report based on the request"""
        try:
            # Create report record
            report = self._create_report_record(request, user_id)
            
            # Update status to generating
            report.status = ReportStatus.GENERATING
            self.db.commit()
            
            # Generate the actual report file
            file_path = self._generate_file(report, request)
            
            # Update report with file information
            report.file_path = str(file_path)
            report.file_name = file_path.name
            report.file_size = file_path.stat().st_size
            report.status = ReportStatus.COMPLETED
            report.generated_at = datetime.utcnow()
            
            # Set expiration date
            expires_in_hours = request.expires_in_hours or 24
            report.expires_at = datetime.utcnow() + timedelta(hours=expires_in_hours)
            
            self.db.commit()
            self.db.refresh(report)
            
            logger.info(f"Report {report.report_number} generated successfully")
            return report
            
        except Exception as e:
            logger.error(f"Error generating report: {e}")
            if 'report' in locals():
                report.status = ReportStatus.FAILED
                report.error_message = str(e)
                self.db.commit()
            raise
    
    def _create_report_record(self, request: ReportGenerationRequest, user_id: int) -> GeneratedReport:
        """Create a new report record"""
        template = self.db.query(ReportTemplate).filter(ReportTemplate.id == request.template_id).first()
        if not template:
            raise ValueError("Report template not found")
        
        report_number = f"RPT-{datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:8].upper()}"
        
        report = GeneratedReport(
            report_number=report_number,
            template_id=request.template_id,
            report_type=template.report_type,
            report_format=request.report_format,
            parameters=request.parameters,
            date_range_start=request.date_range_start,
            date_range_end=request.date_range_end,
            created_by=user_id
        )
        
        self.db.add(report)
        self.db.commit()
        self.db.refresh(report)
        return report
    
    def _generate_file(self, report: GeneratedReport, request: ReportGenerationRequest) -> Path:
        """Generate the actual report file"""
        if request.report_format == ReportFormat.PDF:
            return self._generate_pdf(report, request)
        elif request.report_format == ReportFormat.EXCEL:
            return self._generate_excel(report, request)
        elif request.report_format == ReportFormat.CSV:
            return self._generate_csv(report, request)
        elif request.report_format == ReportFormat.HTML:
            return self._generate_html(report, request)
        else:
            raise ValueError(f"Unsupported report format: {request.report_format}")
    
    def _generate_pdf(self, report: GeneratedReport, request: ReportGenerationRequest) -> Path:
        """Generate PDF report"""
        filename = f"{report.report_number}.pdf"
        file_path = self.temp_dir / filename
        
        doc = SimpleDocTemplate(str(file_path), pagesize=A4)
        story = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Add title
        title = Paragraph(f"Relatório - {report.report_type.value.title()}", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Add report info
        info_data = [
            ['Número do Relatório:', report.report_number],
            ['Tipo:', report.report_type.value.title()],
            ['Data de Geração:', datetime.now().strftime('%d/%m/%Y %H:%M')],
            ['Período:', self._format_date_range(report.date_range_start, report.date_range_end)]
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Add report content based on type
        content = self._get_report_content(report, request)
        story.extend(content)
        
        # Build PDF
        doc.build(story)
        return file_path
    
    def _generate_excel(self, report: GeneratedReport, request: ReportGenerationRequest) -> Path:
        """Generate Excel report"""
        filename = f"{report.report_number}.xlsx"
        file_path = self.temp_dir / filename
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório"
        
        # Add header information
        ws['A1'] = f"Relatório - {report.report_type.value.title()}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A3'] = 'Número do Relatório:'
        ws['B3'] = report.report_number
        ws['A4'] = 'Tipo:'
        ws['B4'] = report.report_type.value.title()
        ws['A5'] = 'Data de Geração:'
        ws['B5'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        ws['A6'] = 'Período:'
        ws['B6'] = self._format_date_range(report.date_range_start, report.date_range_end)
        
        # Add data
        data = self._get_report_data(report, request)
        if data:
            df = pd.DataFrame(data)
            
            # Start data from row 8
            start_row = 8
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Style the header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[start_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(str(file_path))
        return file_path
    
    def _generate_csv(self, report: GeneratedReport, request: ReportGenerationRequest) -> Path:
        """Generate CSV report"""
        filename = f"{report.report_number}.csv"
        file_path = self.temp_dir / filename
        
        data = self._get_report_data(report, request)
        if data:
            df = pd.DataFrame(data)
            df.to_csv(file_path, index=False, encoding='utf-8-sig')
        else:
            # Create empty CSV with headers
            with open(file_path, 'w', encoding='utf-8-sig') as f:
                f.write("Nenhum dado encontrado para os critérios especificados\n")
        
        return file_path
    
    def _generate_html(self, report: GeneratedReport, request: ReportGenerationRequest) -> Path:
        """Generate HTML report"""
        filename = f"{report.report_number}.html"
        file_path = self.temp_dir / filename
        
        html_content = f"""
        <!DOCTYPE html>
        <html lang="pt-BR">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Relatório - {report.report_type.value.title()}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .info-table {{ width: 100%; border-collapse: collapse; margin-bottom: 30px; }}
                .info-table th, .info-table td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                .info-table th {{ background-color: #f2f2f2; }}
                .data-table {{ width: 100%; border-collapse: collapse; }}
                .data-table th, .data-table td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                .data-table th {{ background-color: #366092; color: white; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Relatório - {report.report_type.value.title()}</h1>
            </div>
            
            <table class="info-table">
                <tr><th>Número do Relatório</th><td>{report.report_number}</td></tr>
                <tr><th>Tipo</th><td>{report.report_type.value.title()}</td></tr>
                <tr><th>Data de Geração</th><td>{datetime.now().strftime('%d/%m/%Y %H:%M')}</td></tr>
                <tr><th>Período</th><td>{self._format_date_range(report.date_range_start, report.date_range_end)}</td></tr>
            </table>
            
            {self._get_html_content(report, request)}
        </body>
        </html>
        """
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return file_path
    
    def _get_report_content(self, report: GeneratedReport, request: ReportGenerationRequest) -> List:
        """Get report content for PDF generation"""
        content = []
        styles = getSampleStyleSheet()
        
        # Add content based on report type
        if report.report_type.value == "clinical":
            content.extend(self._get_clinical_content(report, request))
        elif report.report_type.value == "financial":
            content.extend(self._get_financial_content(report, request))
        elif report.report_type.value == "commercial":
            content.extend(self._get_commercial_content(report, request))
        elif report.report_type.value == "administrative":
            content.extend(self._get_administrative_content(report, request))
        
        return content
    
    def _get_report_data(self, report: GeneratedReport, request: ReportGenerationRequest) -> List[Dict[str, Any]]:
        """Get report data for Excel/CSV generation"""
        if report.report_type.value == "clinical":
            return self._get_clinical_data(report, request)
        elif report.report_type.value == "financial":
            return self._get_financial_data(report, request)
        elif report.report_type.value == "commercial":
            return self._get_commercial_data(report, request)
        elif report.report_type.value == "administrative":
            return self._get_administrative_data(report, request)
        
        return []
    
    def _get_html_content(self, report: GeneratedReport, request: ReportGenerationRequest) -> str:
        """Get HTML content for HTML report generation"""
        data = self._get_report_data(report, request)
        
        if not data:
            return "<p>Nenhum dado encontrado para os critérios especificados.</p>"
        
        # Convert data to HTML table
        html = "<table class='data-table'>"
        
        # Add headers
        if data:
            headers = list(data[0].keys())
            html += "<tr>"
            for header in headers:
                html += f"<th>{header}</th>"
            html += "</tr>"
            
            # Add data rows
            for row in data:
                html += "<tr>"
                for header in headers:
                    html += f"<td>{row.get(header, '')}</td>"
                html += "</tr>"
        
        html += "</table>"
        return html
    
    def _format_date_range(self, start_date: Optional[datetime], end_date: Optional[datetime]) -> str:
        """Format date range for display"""
        if start_date and end_date:
            return f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
        elif start_date:
            return f"A partir de {start_date.strftime('%d/%m/%Y')}"
        elif end_date:
            return f"Até {end_date.strftime('%d/%m/%Y')}"
        else:
            return "Todos os períodos"
    
    # Clinical report methods
    def _get_clinical_content(self, report: GeneratedReport, request: ReportGenerationRequest) -> List:
        """Get clinical report content for PDF"""
        content = []
        styles = getSampleStyleSheet()
        
        # Add clinical-specific content
        content.append(Paragraph("Resumo Clínico", styles['Heading2']))
        content.append(Spacer(1, 12))
        
        # This would contain actual clinical data
        content.append(Paragraph("Dados clínicos serão implementados aqui.", styles['Normal']))
        
        return content
    
    def _get_clinical_data(self, report: GeneratedReport, request: ReportGenerationRequest) -> List[Dict[str, Any]]:
        """Get clinical report data"""
        # This would query actual clinical data from the database
        return [
            {"Paciente": "João Silva", "Procedimento": "Consulta", "Data": "01/01/2024", "Status": "Concluído"},
            {"Paciente": "Maria Santos", "Procedimento": "Exame", "Data": "02/01/2024", "Status": "Pendente"}
        ]
    
    # Financial report methods
    def _get_financial_content(self, report: GeneratedReport, request: ReportGenerationRequest) -> List:
        """Get financial report content for PDF"""
        content = []
        styles = getSampleStyleSheet()
        
        content.append(Paragraph("Resumo Financeiro", styles['Heading2']))
        content.append(Spacer(1, 12))
        
        content.append(Paragraph("Dados financeiros serão implementados aqui.", styles['Normal']))
        
        return content
    
    def _get_financial_data(self, report: GeneratedReport, request: ReportGenerationRequest) -> List[Dict[str, Any]]:
        """Get financial report data"""
        return [
            {"Data": "01/01/2024", "Receita": "R$ 1.000,00", "Despesa": "R$ 500,00", "Lucro": "R$ 500,00"},
            {"Data": "02/01/2024", "Receita": "R$ 1.200,00", "Despesa": "R$ 600,00", "Lucro": "R$ 600,00"}
        ]
    
    # Commercial report methods
    def _get_commercial_content(self, report: GeneratedReport, request: ReportGenerationRequest) -> List:
        """Get commercial report content for PDF"""
        content = []
        styles = getSampleStyleSheet()
        
        content.append(Paragraph("Resumo Comercial", styles['Heading2']))
        content.append(Spacer(1, 12))
        
        content.append(Paragraph("Dados comerciais serão implementados aqui.", styles['Normal']))
        
        return content
    
    def _get_commercial_data(self, report: GeneratedReport, request: ReportGenerationRequest) -> List[Dict[str, Any]]:
        """Get commercial report data"""
        return [
            {"Procedimento": "Cirurgia Cardíaca", "Orçamentos": 5, "Contratos": 3, "Receita": "R$ 50.000,00"},
            {"Procedimento": "Consulta", "Orçamentos": 20, "Contratos": 18, "Receita": "R$ 3.600,00"}
        ]
    
    # Administrative report methods
    def _get_administrative_content(self, report: GeneratedReport, request: ReportGenerationRequest) -> List:
        """Get administrative report content for PDF"""
        content = []
        styles = getSampleStyleSheet()
        
        content.append(Paragraph("Resumo Administrativo", styles['Heading2']))
        content.append(Spacer(1, 12))
        
        content.append(Paragraph("Dados administrativos serão implementados aqui.", styles['Normal']))
        
        return content
    
    def _get_administrative_data(self, report: GeneratedReport, request: ReportGenerationRequest) -> List[Dict[str, Any]]:
        """Get administrative report data"""
        return [
            {"Usuário": "Dr. João", "Ações": 25, "Última Atividade": "01/01/2024"},
            {"Usuário": "Enfermeira Maria", "Ações": 15, "Última Atividade": "02/01/2024"}
        ]

class ReportService:
    """Service for managing reports"""
    
    def __init__(self, db: Session):
        self.db = db
        self.generator = ReportGenerator(db)
    
    def generate_report(self, request: ReportGenerationRequest, user_id: int) -> GeneratedReport:
        """Generate a new report"""
        return self.generator.generate_report(request, user_id)
    
    def get_report(self, report_id: int, user_id: int) -> Optional[GeneratedReport]:
        """Get a report by ID"""
        return self.db.query(GeneratedReport).filter(
            GeneratedReport.id == report_id,
            GeneratedReport.created_by == user_id
        ).first()
    
    def get_user_reports(self, user_id: int, skip: int = 0, limit: int = 100) -> List[GeneratedReport]:
        """Get user's reports"""
        return self.db.query(GeneratedReport).filter(
            GeneratedReport.created_by == user_id
        ).offset(skip).limit(limit).all()
    
    def delete_report(self, report_id: int, user_id: int) -> bool:
        """Delete a report"""
        report = self.get_report(report_id, user_id)
        if not report:
            return False
        
        # Delete file if exists
        if report.file_path and os.path.exists(report.file_path):
            try:
                os.remove(report.file_path)
            except Exception as e:
                logger.warning(f"Could not delete file {report.file_path}: {e}")
        
        self.db.delete(report)
        self.db.commit()
        return True
    
    def log_access(self, report_id: int, user_id: int, access_type: str, ip_address: str = None, user_agent: str = None):
        """Log report access"""
        access_log = ReportAccessLog(
            report_id=report_id,
            user_id=user_id,
            access_type=access_type,
            ip_address=ip_address,
            user_agent=user_agent
        )
        self.db.add(access_log)
        
        # Update download count
        report = self.db.query(GeneratedReport).filter(GeneratedReport.id == report_id).first()
        if report and access_type == "download":
            report.download_count += 1
        
        self.db.commit()
    
    def cleanup_expired_reports(self):
        """Clean up expired reports"""
        expired_reports = self.db.query(GeneratedReport).filter(
            GeneratedReport.expires_at < datetime.utcnow(),
            GeneratedReport.status == ReportStatus.COMPLETED
        ).all()
        
        for report in expired_reports:
            # Delete file if exists
            if report.file_path and os.path.exists(report.file_path):
                try:
                    os.remove(report.file_path)
                except Exception as e:
                    logger.warning(f"Could not delete expired file {report.file_path}: {e}")
            
            # Update status
            report.status = ReportStatus.EXPIRED
        
        self.db.commit()
        return len(expired_reports)
